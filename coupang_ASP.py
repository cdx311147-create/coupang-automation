"""
쿠팡 Ads Center 상품 가격 수집기 최종 v4
=========================================
실제 확인된 텍스트 구조:
  르까프 콜리프팅 슈즈 LCF 001, GRAY, 230   ← 상품명(, 색상, 사이즈)
  (174)                                       ← 리뷰수
  ID : 83041880756                            ← 옵션ID
  29,090원                                    ← 가격
  재고량 : 품절                               ← 재고
  아이템 워너 : 선정                          ← 제외
  상품 선택                                   ← 제외
  모든 옵션 선택                              ← 제외

사용법:
  1. pip install selenium openpyxl
  2. python coupang_ads_scraper.py
  3. 브라우저에서 서플라이어 허브 로그인
  4. Ads Center → 광고 만들기 → 수동 상품 설정 탭 + 필터 해제
  5. 상품 목록 보이면 Enter → 자동 수집 시작
"""

import re
import os
import time
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    NoSuchElementException, TimeoutException, StaleElementReferenceException
    
)


# ── 설정 ──────────────────────────────────────────────────────────────────────
OUTPUT_FILE   = "쿠팡_가격현황.xlsx"
SAVE_INTERVAL = 20
PAGE_WAIT     = 3.0
LOAD_TIMEOUT  = 10

LOGIN_URL = (
    "https://xauth.coupang.com/auth/realms/seller/protocol/openid-connect/auth"
    "?response_type=code&client_id=supplier-hub&scope=openid"
    "&state=K3WSaWXeA71e-ovfU7n0SWRBTDtbfbsNUc4AgfVbahQ%3D"
    "&redirect_uri=https://supplier.coupang.com/login/oauth2/code/keycloak"
    "&nonce=MZJ3nt-A5_XOg0p0Q3v9Inlz7nd8nVD4lPbfCxE4bAk"
)
ADS_URL = "https://advertising.coupang.com/marketing/campaign/registration?goalType=SALES"

# 제외할 줄 패턴
SKIP_LINES = {"상품 선택", "모든 옵션 선택"}


# ── 드라이버 ──────────────────────────────────────────────────────────────────

def init_driver() -> webdriver.Chrome:
    opts = Options()
    opts.add_argument("--window-size=1400,900")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    return webdriver.Chrome(options=opts)


def safe_click(driver, element):
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", element)
    time.sleep(0.3)
    driver.execute_script("arguments[0].click();", element)


def switch_to_ads_tab(driver):
    for handle in driver.window_handles:
        driver.switch_to.window(handle)
        if "advertising.coupang.com" in driver.current_url:
            return
    driver.switch_to.window(driver.window_handles[-1])


# ── 로딩 대기 ─────────────────────────────────────────────────────────────────

def wait_for_page_load(driver, prev_text: str) -> bool:
    deadline = time.time() + LOAD_TIMEOUT
    while time.time() < deadline:
        try:
            container    = driver.find_element(By.CSS_SELECTOR, "div.vendor-item-list")
            current_text = container.text.strip()
            if current_text and current_text != prev_text:
                if re.search(r"ID\s*:\s*\d{7,}", current_text):
                    return True
        except (NoSuchElementException, StaleElementReferenceException):
            pass
        time.sleep(0.5)
    return False


def get_current_text(driver) -> str:
    try:
        return driver.find_element(By.CSS_SELECTOR, "div.vendor-item-list").text.strip()
    except NoSuchElementException:
        return ""


# ── 파싱 ─────────────────────────────────────────────────────────────────────

def split_product_name(raw: str) -> tuple[str, str, str]:
    """
    "르까프 콜리프팅 슈즈 LCF 001, GRAY, 230"
    → ("르까프 콜리프팅 슈즈 LCF 001", "GRAY", "230")

    "[Lecaf] 르까프 콜리프팅 슈즈 LCF-001, GRAY, 235"
    → 브랜드 태그 [xxx] 제거 후 동일 처리
    """
    # [브랜드] 태그 제거
    raw = re.sub(r"^\[.*?\]\s*", "", raw).strip()

    parts = [p.strip() for p in raw.split(",")]

    if len(parts) >= 3:
        # 마지막: 사이즈, 그 앞: 색상, 나머지: 상품명
        size         = parts[-1]
        color        = parts[-2]
        product_name = ", ".join(parts[:-2])
    elif len(parts) == 2:
        product_name = parts[0]
        # 숫자 포함이면 사이즈, 아니면 색상
        if re.search(r"\d", parts[1]):
            color = ""
            size  = parts[1]
        else:
            color = parts[1]
            size  = ""
    else:
        product_name = raw
        color        = ""
        size         = ""

    return product_name.strip(), color.strip(), size.strip()


def parse_vendor_item_list(driver) -> list[dict]:
    rows = []
    try:
        container = driver.find_element(By.CSS_SELECTOR, "div.vendor-item-list")
        full_text = container.text.strip()
    except NoSuchElementException:
        return rows

    if not full_text:
        return rows

    # ── 줄 단위로 파싱 ───────────────────────────────────────────────────────
    # 전략: 줄을 순서대로 읽으면서 ID 줄 기준으로 앞뒤를 묶음
    lines = [l.strip() for l in full_text.splitlines() if l.strip()]

    # ID 줄 위치 전부 찾기
    id_indices = [i for i, l in enumerate(lines) if re.match(r"ID\s*:\s*\d{7,}", l)]

    for pos, id_idx in enumerate(id_indices):
        # 이 ID 블록의 끝: 다음 ID 줄 직전 또는 끝
        end_idx = id_indices[pos + 1] if pos + 1 < len(id_indices) else len(lines)

        # ID 위쪽 줄들 (상품명, 리뷰수)
        # 직전 ID 블록 끝부터 현재 ID 줄 사이
        prev_end = id_indices[pos - 1] if pos > 0 else 0
        # 직전 블록이면 아이템 워너 / 상품 선택 등 제거 후 상품명 찾기
        before_lines = lines[prev_end:id_idx]

        # 제외 줄 필터
        before_lines = [
            l for l in before_lines
            if l not in SKIP_LINES
            and not l.startswith("아이템 워너")
            and not re.match(r"^\d+,?\d*원$", l)      # 이전 블록 가격 줄 제외
            and not re.match(r"^재고량", l)            # 이전 블록 재고 줄 제외
        ]

        # 리뷰수: (숫자) 패턴
        review      = ""
        raw_name    = ""
        for l in reversed(before_lines):
            if re.match(r"^\(\d[\d,]*\)$", l):
                review = l.strip("()")
            elif not raw_name and not re.match(r"^[\d,\.\s]+$", l):
                raw_name = l

        # 옵션ID
        id_line    = lines[id_idx]
        id_match   = re.search(r"ID\s*:\s*(\d{7,})", id_line)
        option_id  = id_match.group(1) if id_match else ""

        # ID 아래 줄들 (가격, 재고)
        after_lines = lines[id_idx + 1: end_idx]
        after_lines = [
            l for l in after_lines
            if l not in SKIP_LINES and not l.startswith("아이템 워너")
        ]

        # 가격
        price = 0
        for l in after_lines:
            m = re.match(r"^([\d,]+)원$", l)
            if m:
                price = int(m.group(1).replace(",", ""))
                break

        # 재고량
        stock = ""
        for l in after_lines:
            m = re.match(r"^재고량\s*:\s*(.+)$", l)
            if m:
                stock = m.group(1).strip()
                break

        # 상품명, 색상, 사이즈 분리
        product_name, color, size = split_product_name(raw_name) if raw_name else ("", "", "")

        if option_id:
            rows.append({
                "상품명":   product_name,
                "색상":     color,
                "사이즈":   size,
                "리뷰수":   review,
                "옵션ID":   option_id,
                "가격(원)": price,
                "재고량":   stock,
            })

    return rows


# ── 페이지네이션 ──────────────────────────────────────────────────────────────

def get_total_pages(driver) -> int:
    try:
        items = driver.find_elements(
            By.CSS_SELECTOR,
            "div.vendor-item-pagination li[class*='ant-pagination-item-']"
        )
        nums = []
        for el in items:
            m = re.search(r"ant-pagination-item-(\d+)", el.get_attribute("class") or "")
            if m:
                nums.append(int(m.group(1)))
        if nums:
            return max(nums)
        pagination = driver.find_element(By.CSS_SELECTOR, "div.vendor-item-pagination")
        nums = [int(n) for n in re.findall(r"\b(\d+)\b", pagination.text) if int(n) > 1]
        return max(nums) if nums else 1
    except Exception:
        return 1


def go_next_page(driver, current_page: int) -> bool:
    try:
        pagination = driver.find_element(By.CSS_SELECTOR, "div.vendor-item-pagination")
        btns = pagination.find_elements(
            By.CSS_SELECTOR, f"li.ant-pagination-item-{current_page + 1}"
        )
        if btns and btns[0].is_displayed():
            safe_click(driver, btns[0])
            return True
        next_btn = pagination.find_element(By.CSS_SELECTOR, "li.ant-pagination-next")
        if next_btn.get_attribute("aria-disabled") == "true":
            return False
        safe_click(driver, next_btn)
        return True
    except NoSuchElementException:
        return False
    except Exception as e:
        print(f"\n    [페이지 이동 오류] {e}")
        return False


# ── 엑셀 저장 ─────────────────────────────────────────────────────────────────

def save_excel(rows: list[dict]) -> str:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "쿠팡 가격 현황"

    cols = ["상품명", "색상", "사이즈", "리뷰수", "옵션ID", "가격(원)", "재고량"]
    hf   = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
    hfil = PatternFill("solid", start_color="C00000")
    ca   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    la   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin = Side(style="thin")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[1].height = 28
    for ci, cn in enumerate(cols, 1):
        c = ws.cell(1, ci, cn)
        c.font = hf; c.fill = hfil; c.alignment = ca; c.border = bdr

    af = PatternFill("solid", start_color="FFF2F2")
    for ri, row in enumerate(rows, 2):
        for ci, cn in enumerate(cols, 1):
            v = row.get(cn, "")
            c = ws.cell(ri, ci, v)
            c.font = Font(name="맑은 고딕", size=10); c.border = bdr
            if ri % 2 == 0: c.fill = af
            if cn == "가격(원)" and isinstance(v, int):
                c.number_format = '#,##0"원"'
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif cn == "상품명": c.alignment = la
            else: c.alignment = ca

    for ci, w in {1: 45, 2: 12, 3: 15, 4: 10, 5: 15, 6: 15, 7: 12}.items():
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws2 = wb.create_sheet("통계")
    ws2["A1"] = "최종 수집 일시"; ws2["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws2["A2"] = "총 상품 수";     ws2["B2"] = len(rows)

    wb.save(OUTPUT_FILE)
    return OUTPUT_FILE


# ── 메인 ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  쿠팡 Ads Center 상품 가격 수집기 v4")
    print("=" * 60)
    print(f"  저장 파일: {OUTPUT_FILE} (20페이지마다 덮어씀)")

    try:
        import openpyxl
    except ImportError:
        print("[설치 필요] pip install selenium openpyxl")
        return

    driver = init_driver()

    try:
        # Step 1: 로그인
        print("\n[Step 1] 서플라이어 허브 로그인 페이지 오픈")
        driver.get(LOGIN_URL)
        input("\n  ✋ 로그인 완료 후 Enter 누르세요... ")

        # Step 2: Ads Center 이동
        print("\n[Step 2] Ads Center 이동 중...")
        driver.get(ADS_URL)
        time.sleep(3)
        switch_to_ads_tab(driver)
        print(f"  현재 URL: {driver.current_url}")

        # Step 3: 수동 설정 안내
        print("\n[Step 3] 아래 순서로 직접 진행해주세요:")
        print("  1) 페이지 스크롤 내려서 '광고 설정' 섹션 확인")
        print("  2) '수동 상품 설정' 탭 클릭")
        print("  3) 상단 필터 체크박스 모두 해제")
        input("\n  ✋ 상품 목록이 보이는 상태에서 Enter 누르세요... ")

        switch_to_ads_tab(driver)

        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div.vendor-item-list"))
            )
            print("  ✅ 상품 목록 확인됨")
        except TimeoutException:
            print(f"  ❌ 상품 목록 없음 (현재 URL: {driver.current_url})")
            input("  ✋ 상품 목록이 보이는 상태에서 Enter... ")

        # Step 4: 수집
        total_pages = get_total_pages(driver)
        print(f"\n[Step 4] 수집 시작 — 총 {total_pages:,}페이지 (약 {total_pages*5:,}개 상품)\n")

        all_rows: list[dict] = []
        seen_ids: set        = set()
        prev_text            = ""

        for page in range(1, total_pages + 1):
            print(f"  [{page:>4}/{total_pages}] 로딩 대기...", end=" ", flush=True)

            loaded = wait_for_page_load(driver, prev_text)
            if not loaded:
                print("⚠️  타임아웃, 재시도", end=" ")
                time.sleep(2)

            prev_text = get_current_text(driver)
            rows      = parse_vendor_item_list(driver)
            new_rows  = [r for r in rows if r["옵션ID"] not in seen_ids]
            for r in new_rows:
                seen_ids.add(r["옵션ID"])
            all_rows.extend(new_rows)

            if new_rows:
                print(f"✅ {len(new_rows)}개 (누적 {len(all_rows):,}개)")
                for r in new_rows:
                    price_str = f"{r['가격(원)']:,}원" if r["가격(원)"] else "가격미확인"
                    name      = f"{r['상품명']} / {r['색상']} / {r['사이즈']}".strip(" /")
                    name      = name[:50] + ("…" if len(name) > 50 else "")
                    print(f"       [{r['옵션ID']}] {name} — {price_str}  재고:{r['재고량']}")
            else:
                print("⬜ 0개")

            if page % SAVE_INTERVAL == 0 and all_rows:
                save_excel(all_rows)
                print(f"\n  💾 [{page}p 저장] '{OUTPUT_FILE}' ({len(all_rows):,}개)\n")

            if page < total_pages:
                if not go_next_page(driver, page):
                    print(f"\n  ⚠️  {page+1}페이지 이동 실패 — 종료")
                    break

        # Step 5: 최종 저장
        print("\n" + "=" * 60)
        if all_rows:
            save_excel(all_rows)
            print(f"✅ [완료] '{OUTPUT_FILE}'")
            print(f"   총 {len(all_rows):,}개 상품 수집")
        else:
            print("❌ 수집된 상품이 없습니다.")
        print("=" * 60)

    finally:
        input("\n  브라우저를 닫으려면 Enter... ")
        driver.quit()


if __name__ == "__main__":
    main()