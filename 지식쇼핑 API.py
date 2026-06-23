import os
import requests
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

# 1. 네이버 API 키 설정
client_id = "z0EdqFGImR8zOnr6wchK"
client_secret = "9417zWeUn1"

# 2. 경로 설정
base_path = os.path.dirname(os.path.abspath(__file__))
input_file = os.path.join(base_path, "검색 상품명.txt")
output_file = os.path.join(base_path, f"ASP_전수조사_리포트_{datetime.now().strftime('%m%d_%H%M')}.xlsx")

def get_shopping_data(query):
    url = "https://openapi.naver.com/v1/search/shop.json"
    headers = {"X-Naver-Client-Id": client_id, "X-Naver-Client-Secret": client_secret}
    params = {"query": query, "display": 100, "sort": "asc"} # 최저가순 100개
    try:
        response = requests.get(url, headers=headers, params=params)
        return response.json().get('items', []) if response.status_code == 200 else []
    except: return []

def main():
    if not os.path.exists(input_file): return

    target_list = []
    with open(input_file, 'r', encoding='utf-8') as f:
        for idx, line in enumerate(f):
            if ',' in line:
                name, price = line.strip().split(',')
                target_list.append({"order": idx, "name": name, "base_price": int(price)})

    all_raw_data = []
    for target in target_list:
        print(f"전수 조사 중 (필터 없음): {target['name']}")
        items = get_shopping_data(target['name'])
        
        for item in items:
            cur_price = int(item['lprice'])
            real_title = item['title'].replace('<b>', '').replace('</b>', '')
            
            # [수정 로직] 유사도/매칭 로직 완전 삭제! 
            # 오직 '기준가보다 낮은가'만 확인하고 다 가져옵니다.
            if cur_price < target['base_price']:
                all_raw_data.append([
                    target['order'],
                    target['name'],
                    target['base_price'],
                    cur_price,
                    target['base_price'] - cur_price,
                    item['mallName'],
                    real_title
                ])

    if not all_raw_data:
        print("✅ 기준가 미달 상품이 없습니다. 완벽한 방어 중입니다!")
        return

    df = pd.DataFrame(all_raw_data, columns=['순서', '검색어', '기준가', '현재최저가', '하락액', '판매처', '실제상품명'])
    df = df.sort_values(by=['순서', '현재최저가'], ascending=[True, True])

    # --- [디자인: 필터 걸기 좋은 엑셀 구조] ---
    workbook = Workbook()
    ws = workbook.active
    ws.title = "ASP_전수검사_결과"
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    item_fill = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
    border_all = Border(left=Side(style='thin', color="BBBBBB"), right=Side(style='thin', color="BBBBBB"), 
                        top=Side(style='thin', color="BBBBBB"), bottom=Side(style='thin', color="BBBBBB"))

    headers = ["검색어", "기준가", "현재최저가", "하락액", "판매처", "실제상품명"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=1, column=i+1)
        cell.value = h
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_all

    curr_row = 2
    for name, group in df.groupby('검색어', sort=False):
        start_row = curr_row
        for _, row in group.iterrows():
            ws.cell(row=curr_row, column=3).value = f"{row['현재최저가']:,}원"
            ws.cell(row=curr_row, column=4).value = f"-{row['하락액']:,}원"
            ws.cell(row=curr_row, column=4).font = Font(color="C62828", bold=True)
            ws.cell(row=curr_row, column=5).value = row['판매처']
            ws.cell(row=curr_row, column=6).value = row['실제상품명']
            
            for c in range(3, 7):
                ws.cell(row=curr_row, column=c).border = border_all
                ws.cell(row=curr_row, column=c).alignment = Alignment(vertical="center", wrap_text=True)
            curr_row += 1
        
        # 검색어/기준가 열 병합 및 스타일 (메모장 순서 유지)
        for col_idx, val in enumerate([name, f"{group['기준가'].iloc[0]:,}원"], 1):
            ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=curr_row-1, end_column=col_idx)
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = val
            cell.fill = item_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True)
            for r in range(start_row, curr_row):
                ws.cell(row=r, column=col_idx).border = border_all

    # 레이아웃 최적화
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 50
    
    ws.freeze_panes = "A2"
    workbook.save(output_file)
    print(f"\n✅ 필터 없는 전수조사 완료! (결과 파일: {os.path.basename(output_file)})")

if __name__ == "__main__":
    main()