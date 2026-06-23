"""
네이버 쇼핑 오퍼리스트 자동화 스크립트 v5
1단계 - 네이버 API: 쿠팡/크림/리셀/해외 1차 필터
2단계 - Selenium: 링크 직접 접속 → 판매자명 확인 → 특정 판매처 제거
"""

import os
import re
import time
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# =============================================
# ★ 여기에 네이버 API 키를 입력하세요 ★
NAVER_CLIENT_ID = "z0EdqFGImR8zOnr6wchK"
NAVER_CLIENT_SECRET = "9417zWeUn1"
# =============================================

# 1단계: API 레벨 제외 판매처
EXCLUDE_SELLERS_API = [
    "쿠팡", "coupang",
    "크림", "kream",
    "솔드아웃", "soldout",
    "무신사 스니커즈",
]

# 2단계: 크롤링 레벨 제외 판매처 (페이지 내 판매자명 기준)
EXCLUDE_SELLERS_CRAWL = [
    "구하다",
    "어도어럭스", "adorelux",
    "비뉴",
    "밀리언캔디",
    # 1단계에서 못 걸러진 경우 대비 중복 추가
    "쿠팡", "크림", "kream",
]

# 해외 키워드 — 페이지 내 텍스트에서 발견 시 제외
OVERSEAS_KEYWORDS = [
    "해외배송", "해외직구", "해외구매", "구매대행", "병행수입",
]

# Selenium 페이지 로딩 대기 시간 (초)
PAGE_LOAD_WAIT = 6


def load_products(filepath="item_list.txt"):
    if not os.path.exists(filepath):
        print(f"[오류] '{filepath}' 파일을 찾을 수 없습니다.")
        return []
    with open(filepath, "r", encoding="utf-8") as f:
        lines = [line.strip() for line in f.readlines()]
    products = [l for l in lines if l and not l.startswith("#")]
    print(f"[정보] 상품 {len(products)}개 로드 완료\n")
    return products


def clean_html(text):
    return re.sub(r"<[^>]+>", "", text).strip()


def parse_price(price_str):
    try:
        return int(re.sub(r"[^\d]", "", str(price_str)))
    except:
        return 0


def init_selenium():
    """Selenium Chrome 드라이버 초기화 (헤드리스)"""
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
    driver = webdriver.Chrome(options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver


def search_naver_shopping(query):
    """네이버 쇼핑 API 검색 (낮은 가격순)"""
    url = "https://openapi.naver.com/v1/search/shop.json"
    headers = {
        "X-Naver-Client-Id": NAVER_CLIENT_ID,
        "X-Naver-Client-Secret": NAVER_CLIENT_SECRET,
    }
    params = {"query": query, "display": 100, "sort": "asc"}
    try:
        resp = requests.get(url, headers=headers, params=params, timeout=10)
        if resp.status_code == 200:
            return resp.json().get("items", [])
        else:
            print(f"  [API 오류] {resp.status_code}")
            return []
    except Exception as e:
        print(f"  [API 요청 실패] {e}")
        return []


def api_filter(items, product_name):
    """
    1단계: API 응답에서 쿠팡/크림/리셀/해외 제거
    nation, productType, 판매처명 키워드로 필터
    """
    results = []
    for item in items:
        mall_name = clean_html(item.get("mallName", ""))
        price = parse_price(item.get("lprice", 0))
        link = item.get("link", "")

        if price <= 0:
            continue

        # 판매처명 키워드 필터
        mall_lower = mall_name.lower()
        if any(kw in mall_lower for kw in EXCLUDE_SELLERS_API):
            continue

        # productType 6 = 해외상품
        if str(item.get("productType", "")) == "6":
            continue

        # nation 필드
        nation = item.get("nation", "").upper()
        if nation and nation != "KR":
            continue

        results.append({
            "상품명": product_name,
            "판매처": mall_name,
            "판매가": price,
            "링크": link,
        })

    print(f"  [1단계 API] {len(items)}개 → {len(results)}개 통과")
    return results


def crawl_and_verify(driver, item):
    """
    2단계: Selenium으로 링크 접속 → 판매자명/해외 여부 확인
    반환: True(통과) / False(제외)
    """
    link = item["링크"]
    try:
        driver.get(link)
        time.sleep(PAGE_LOAD_WAIT)
        page_text = driver.find_element(By.TAG_NAME, "body").text.lower()

        # 해외 키워드 확인
        for kw in OVERSEAS_KEYWORDS:
            if kw in page_text:
                print(f"    [제외-해외] '{kw}' 키워드 발견")
                return False

        # 제외 판매처 확인
        for kw in EXCLUDE_SELLERS_CRAWL:
            if kw.lower() in page_text:
                print(f"    [제외-판매처] '{kw}' 발견")
                return False

        return True

    except Exception as e:
        print(f"    [크롤링 오류] {e}")
        # 오류 시 통과 처리 (보수적으로 제외하려면 False로 변경)
        return True


def save_to_excel(all_results, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "오퍼리스트"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="2E4057")
    header_align = Alignment(horizontal="center", vertical="center")
    data_font = Font(name="Arial", size=10)
    alt_fill = PatternFill("solid", start_color="F2F4F7")
    no_result_font = Font(name="Arial", size=10, color="AAAAAA", italic=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["상품명", "판매처", "판매가", "링크"]
    col_widths = [38, 22, 14, 15]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"

    for row_idx, row in enumerate(all_results, 2):
        is_alt = (row_idx % 2 == 1)
        no_result = row.get("no_result", False)

        for col_idx in range(1, 5):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border

            if col_idx == 1:
                cell.value = row["상품명"]
                cell.font = data_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if is_alt and not no_result:
                    cell.fill = alt_fill

            elif col_idx == 2:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if no_result:
                    cell.value = "검색 결과 없음"
                    cell.font = no_result_font
                else:
                    cell.value = row["판매처"]
                    cell.font = data_font
                    if is_alt:
                        cell.fill = alt_fill

            elif col_idx == 3:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if no_result:
                    cell.value = "-"
                    cell.font = no_result_font
                else:
                    cell.value = row["판매가"]
                    cell.font = data_font
                    cell.number_format = "#,##0"
                    if is_alt:
                        cell.fill = alt_fill

            elif col_idx == 4:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if no_result:
                    cell.value = "-"
                    cell.font = no_result_font
                else:
                    cell.hyperlink = row["링크"]
                    cell.value = "링크 열기"
                    cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

        ws.row_dimensions[row_idx].height = 18

    wb.save(output_path)
    found = sum(1 for r in all_results if not r.get("no_result"))
    no_res = len(all_results) - found
    print(f"\n[완료] 저장: {output_path}")
    print(f"[완료] 전체 {len(all_results)}개 | 결과있음: {found}개 | 결과없음: {no_res}개")


def main():
    print("=" * 55)
    print("  네이버 쇼핑 오퍼리스트 자동화 v5 (Selenium)")
    print("=" * 55)

    if "여기에" in NAVER_CLIENT_ID:
        print("[오류] 스크립트 상단에 네이버 API 키를 입력하세요!")
        return

    products = load_products("item_list.txt")
    if not products:
        return

    # Selenium 초기화 (한 번만)
    print("[정보] Selenium 브라우저 초기화 중...")
    driver = init_selenium()
    print("[정보] 브라우저 준비 완료\n")

    all_results = []

    try:
        for i, product in enumerate(products, 1):
            print(f"[{i}/{len(products)}] {product}")

            # 1단계: API 필터
            items = search_naver_shopping(product)
            if not items:
                print(f"  → API 검색 결과 없음")
                all_results.append({"상품명": product, "no_result": True})
                continue

            candidates = api_filter(items, product)
            if not candidates:
                print(f"  → 1단계 필터 후 결과 없음")
                all_results.append({"상품명": product, "no_result": True})
                continue

            # 2단계: Selenium 크롤링 — 가격 낮은 순으로 하나씩 확인, 통과하면 바로 채택
            print(f"  [2단계 크롤링] {len(candidates)}개 링크 순서대로 확인...")
            best = None
            for idx, candidate in enumerate(candidates, 1):
                print(f"    ({idx}/{len(candidates)}) {candidate['판매처']} {candidate['판매가']:,}원")
                if crawl_and_verify(driver, candidate):
                    best = candidate
                    print(f"    → ✅ 통과! 최저가 채택")
                    break  # 가격 낮은 순이므로 첫 번째 통과가 최저가
                else:
                    print(f"    → ❌ 제외")

            if best:
                print(f"  → 최종: {best['판매처']} {best['판매가']:,}원")
                all_results.append(best)
            else:
                print(f"  → 모든 링크 제외됨 → 검색 결과 없음")
                all_results.append({"상품명": product, "no_result": True})

    finally:
        driver.quit()
        print("\n[정보] 브라우저 종료")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"오퍼리스트_결과_{timestamp}.xlsx"
    save_to_excel(all_results, output_path)
    print("\n작업 완료!")


if __name__ == "__main__":
    main()